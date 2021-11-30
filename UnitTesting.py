#  pip install pyinstaller
#  pyinstaller yourprogram.py
#  pyinstaller -F yourprogram.py

import pandas as pd
from easygui import fileopenbox
import openpyxl
from openpyxl import load_workbook, Workbook
from easygui import fileopenbox
from datetime import datetime, timedelta, time
import numpy as np
import csv
import os

fn = fileopenbox()
#fn = 'C:\\Users\\andre\\Documents\\School\\Fall2021Semester\\ECE 591\\RTS-Calibration-Verification\\data\\11-2-21-CLEAN.xlsx'
if(fn.endswith('.csv')):
    wb = Workbook()
    ws = wb.active
    with open(fn) as f:
        reader = csv.reader(f, delimiter=',')
        for row_index, row in enumerate(reader):
            for column_index, cell in enumerate(row):
                column_letter = openpyxl.utils.get_column_letter((column_index + 1))
                s = cell
                #Handles heading row or non floats
                try:
                    s = float(s)
                    ws[('%s%s'%(column_letter, (row_index + 1)))].value = s
                    ws[('%s%s'%(column_letter, (row_index + 1)))].number_format = ""
                except ValueError:
                    ws[('%s%s'%(column_letter, (row_index + 1)))].value = cell
    temppath = os.path.dirname(os.path.realpath(__file__)) + '\\data\\.DataCopy.xlsx'
    wb.save(temppath)
else:
    wb = load_workbook(fn)

ws1 = wb.worksheets[0]