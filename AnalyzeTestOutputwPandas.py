# Install c++ runtime environment if not present -- latest version here: https://docs.microsoft.com/en-us/cpp/windows/latest-supported-vc-redist?view=msvc-160
# pip install numpy
# pip install pandas
# pip install easygui
# pip install openpyxl


#  pip install pyinstaller
#  pyinstaller yourprogram.py
#  pyinstaller -F yourprogram.py
import pandas as pd
from easygui import fileopenbox
import openpyxl
from openpyxl import load_workbook, Workbook
from easygui import fileopenbox
from datetime import datetime, timedelta, time
import numpy as np
import csv
import os

def CellIdForValueInColumn(val, col):
    rowId = 1
    cellId = col + str(rowId)
    focusedCell = ws1[cellId]
    while(focusedCell.value != None and val not in str(focusedCell.value)):
        rowId = rowId + 1
        cellId = col+ str(rowId)
        focusedCell = ws1[cellId]
    return col,rowId

fn = fileopenbox()
#fn = 'C:\\Users\\andre\\Documents\\School\\Fall2021Semester\\ECE 591\\RTS-Calibration-Verification\\data\\11-2-21-CLEAN.xlsx'
if(fn.endswith('.csv')):
    wb = Workbook()
    ws = wb.active
    with open(fn) as f:
        reader = csv.reader(f, delimiter=',')
        for row_index, row in enumerate(reader):
            for column_index, cell in enumerate(row):
                column_letter = openpyxl.utils.get_column_letter((column_index + 1))
                s = cell
                #Handles heading row or non floats
                try:
                    s = float(s)
                    ws[('%s%s'%(column_letter, (row_index + 1)))].value = s

                except ValueError:
                    ws[('%s%s'%(column_letter, (row_index + 1)))].value = cell
    temppath = os.path.dirname(os.path.realpath(__file__)) + '\\data\\.DataCopy.xlsx'
    wb.save(temppath)
else:
    wb = load_workbook(fn)

ws1 = wb.worksheets[0]

rowId = 1
cellId = 'A'+ str(rowId)
focusedCell = ws1[cellId]

col,row = CellIdForValueInColumn("Record", 'A')

if(fn.endswith('.csv')):
    df = pd.read_excel(temppath, skiprows=(row-1))
else:
    df = pd.read_excel(fn, skiprows=(row-1))
df.columns = df.columns.str.replace("\"","")
df.columns = df.columns.str.replace(" ","")
cols = df.columns

phaseTests = {
    "LVPT": [["P5V4to6", "P5V1to3", 5],
             ["P50V4to6", "P50V1to3", 50],
             ["P100V4to6", "P100V1to3", 100], 
             ["P150V4to6", "P150V1to3", 150] ],

    "LIPT": [["P1A4to6", "P1A1to3", 1],
             ["P3A4to6", "P3A1to3", 3],
             ["P6A4to6", "P6A1to3", 6] ],

    "HVPT": [["P5HV", 5],
             ["P50HV", 50],
             ["P100HV", 100], 
             ["P150HV", 150] ],

    "HIPT": [["P1HA", 1],
             ["P3HA", 3],
             ["P6HA", 6] ],
}

expectedVals = {
    "VA": [25,50,75,100,125,150,175,200,225,250],
    "IA": [1,2,3,4,5,6],
    "PFDA": [0, 60, 120, 180, -120, -60],
    "VH": [50,100,150,200,250,300],
    "IH": [1,2,3,4,5,6],
}

startbit2 = [[15,10,5 ], [1.5,1,5]]
startbitt2 = [[15,10,15] , [1.5,1,1.5]]

tests = []
testcols = ["TYPE","EXPECTEDVAL",
            "V1", "V2", "V3", "V4", "V5", "V6",
            "Phase(V1)", "Phase(V2)", "Phase(V3)", "Phase(V4)", "Phase(V5)", "Phase(V6)",
            "I1", "I2", "I3", "I4", "I5", "I6",
            "Phase(I1)", "Phase(I2)", "Phase(I3)", "Phase(I4)", "Phase(I5)", "Phase(I6)",

            "VH1", "VH2", "VH3",
            "Phase(VH1)", "Phase(VH2)", "Phase(VH3)",
            "IH1", "IH2", "IH3",
            "Phase(IH1)", "Phase(IH2)", "Phase(IH3)"]

def CompareWithTolerance(expectedVal, val, tolerance):
    decimalPercent = tolerance / 200.0
    highRange = expectedVal * (1.0 + decimalPercent)
    lowRange = expectedVal * (1.0 - decimalPercent)
    if(lowRange <= val and val <= highRange):
        return True
    else:
        return False    
def toPhase(PF):
    return (np.arccos(PF) * 180.0) / np.pi

v1to3pop = False
# I can't believe I need to do this...
def setv1to3True():
    global v1to3pop
    v1to3pop = True
def setv1to3False():
    global v1to3pop
    v1to3pop = False

i1to3pop = False
# I still can't believe I need to do this...
def seti1to3True():
    global i1to3pop
    i1to3pop = True
def seti1to3False():
    global i1to3pop
    i1to3pop = False

i4to6pop = False
# I STILL can't believe I need to do this...
def seti4to6True():
    global i4to6pop
    i4to6pop = True
def seti4to6False():
    global i4to6pop
    i4to6pop = False

vlowcomplete = False
# I STILL can't believe I need to do this...
def setvlowcompletetrue():
    global vlowcomplete
    vlowcomplete = True
def setvlowcompleteFalse():
    global vlowcomplete
    vlowcomplete = False

def FilterAndIdentify(row):
    if(vlowcomplete != True):
        for expectedVal in expectedVals["VA"]:
            if(CompareWithTolerance(expectedVal, row["VA"], 1) and
                CompareWithTolerance(expectedVal, row["VB"], 1) and
                CompareWithTolerance(expectedVal, row["VC"], 1) and
                CompareWithTolerance(0, row["IA"], 1) and
                CompareWithTolerance(0, row["IB"], 1) and
                CompareWithTolerance(0, row["IC"], 1) and
                CompareWithTolerance(1, row["PFDA"], .1) and
                CompareWithTolerance(1, row["PFDB"], .1) and
                CompareWithTolerance(1, row["PFDC"], .1)
                ):
                if(v1to3pop):
                    tests.append(["V4to6",expectedVal, 0,0,0,row["VA"], row["VB"],row["VC"],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0])
                else:
                    tests.append(["V1to3",expectedVal,row["VA"], row["VB"],row["VC"],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0])

        for expectedVal in expectedVals["PFDA"]:
            if(CompareWithTolerance(5, row["VA"], 1) and
                CompareWithTolerance(5, row["VB"], 1) and
                CompareWithTolerance(5, row["VC"], 1) and
                CompareWithTolerance(1, row["IA"], 1) and
                CompareWithTolerance(1, row["IB"], 1) and
                CompareWithTolerance(1, row["IC"], 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDA"]), 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDB"]), 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDC"]), 1)
                ):
                if(v1to3pop):
                    tests.append(["P5V4to6",expectedVal,0,0,0,0,0,0,0,0,0,toPhase(row["PFDA"]), toPhase(row["PFDB"]),toPhase(row["PFDC"]),0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0])
                else:
                    tests.append(["P5V1to3",expectedVal,0,0,0,0,0,0,toPhase(row["PFDA"]), toPhase(row["PFDB"]),toPhase(row["PFDC"]),0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0])

            elif(CompareWithTolerance(50, row["VA"], 1) and
                CompareWithTolerance(50, row["VB"], 1) and
                CompareWithTolerance(50, row["VC"], 1) and
                CompareWithTolerance(1, row["IA"], 1) and
                CompareWithTolerance(1, row["IB"], 1) and
                CompareWithTolerance(1, row["IC"], 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDA"]), 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDB"]), 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDC"]), 1)
                ):
                if(v1to3pop):
                    tests.append(["P50V4to6",expectedVal,0,0,0,0,0,0,0,0,0,toPhase(row["PFDA"]), toPhase(row["PFDB"]),toPhase(row["PFDC"]),0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0])
                else:
                    tests.append(["P50V1to3",expectedVal,0,0,0,0,0,0,toPhase(row["PFDA"]), toPhase(row["PFDB"]),toPhase(row["PFDC"]),0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0])

            elif(CompareWithTolerance(100, row["VA"], 1) and
                CompareWithTolerance(100, row["VB"], 1) and
                CompareWithTolerance(100, row["VC"], 1) and
                CompareWithTolerance(1, row["IA"], 1) and
                CompareWithTolerance(1, row["IB"], 1) and
                CompareWithTolerance(1, row["IC"], 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDA"]), 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDB"]), 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDC"]), 1)
                ):
                if(v1to3pop):
                    tests.append(["P100V4to6",expectedVal,0,0,0,0,0,0,0,0,0,toPhase(row["PFDA"]), toPhase(row["PFDB"]),toPhase(row["PFDC"]),0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0])
                else:
                    tests.append(["P100V1to3",expectedVal,0,0,0,0,0,0,toPhase(row["PFDA"]), toPhase(row["PFDB"]),toPhase(row["PFDC"]),0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0])

            elif(CompareWithTolerance(150, row["VA"], 1) and
                CompareWithTolerance(150, row["VB"], 1) and
                CompareWithTolerance(150, row["VC"], 1) and
                CompareWithTolerance(1, row["IA"], 1) and
                CompareWithTolerance(1, row["IB"], 1) and
                CompareWithTolerance(1, row["IC"], 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDA"]), 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDB"]), 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDC"]), 1)
                ):
                if(v1to3pop):
                    tests.append(["P150V4to6", expectedVal,0,0,0,0,0,0,0,0,0,toPhase(row["PFDA"]), toPhase(row["PFDB"]),toPhase(row["PFDC"]),0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0])
                else:
                    tests.append(["P150V1to3",expectedVal,0,0,0,0,0,0,toPhase(row["PFDA"]), toPhase(row["PFDB"]),toPhase(row["PFDC"]),0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0])
                if(CompareWithTolerance(180, toPhase(row["PFDA"]), 1)):
                    setv1to3True()

        for expectedVal in expectedVals["IA"]:
            if( CompareWithTolerance(0, row["VA"], 1) and
                CompareWithTolerance(0, row["VB"], 1) and
                CompareWithTolerance(0, row["VC"], 1) and
                CompareWithTolerance(expectedVal, row["IA"], 1) and
                CompareWithTolerance(expectedVal, row["IB"], 1) and
                CompareWithTolerance(expectedVal, row["IC"], 1) and
                CompareWithTolerance(1, row["PFDA"], .1) and
                CompareWithTolerance(1, row["PFDB"], .1) and
                CompareWithTolerance(1, row["PFDC"], .1)
                ):
                if(i1to3pop):
                    tests.append(["I4to6",expectedVal, 0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,row["IA"], row["IB"],row["IC"],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0])
                else:
                    tests.append(["I1to3",expectedVal, 0,0,0,0,0,0,0,0,0,0,0,0,row["IA"], row["IB"],row["IC"],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0])
                
        for expectedVal in expectedVals["PFDA"]:
            if(CompareWithTolerance(10, row["VA"], 1) and
                CompareWithTolerance(10, row["VB"], 1) and
                CompareWithTolerance(10, row["VC"], 1) and
                CompareWithTolerance(1, row["IA"], 1) and
                CompareWithTolerance(1, row["IB"], 1) and
                CompareWithTolerance(1, row["IC"], 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDA"]), 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDB"]), 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDC"]), 1)
                ):
                if(i1to3pop):
                    tests.append(["P1A4to6",expectedVal, 0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,toPhase(row["PFDA"]), toPhase(row["PFDB"]),toPhase(row["PFDC"]),0,0,0,0,0,0,0,0,0,0,0,0])
                else:
                    tests.append(["P1A1to3",expectedVal,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,toPhase(row["PFDA"]), toPhase(row["PFDB"]),toPhase(row["PFDC"]),0,0,0,0,0,0,0,0,0,0,0,0,0,0,0])

            elif(CompareWithTolerance(10, row["VA"], 1) and
                CompareWithTolerance(10, row["VB"], 1) and
                CompareWithTolerance(10, row["VC"], 1) and
                CompareWithTolerance(3, row["IA"], 1) and
                CompareWithTolerance(3, row["IB"], 1) and
                CompareWithTolerance(3, row["IC"], 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDA"]), 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDB"]), 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDC"]), 1)
                ):
                if(i1to3pop):
                    tests.append(["P3A4to6",expectedVal, 0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,toPhase(row["PFDA"]), toPhase(row["PFDB"]),toPhase(row["PFDC"]),0,0,0,0,0,0,0,0,0,0,0,0])
                else:
                    tests.append(["P3A1to3",expectedVal,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,toPhase(row["PFDA"]), toPhase(row["PFDB"]),toPhase(row["PFDC"]),0,0,0,0,0,0,0,0,0,0,0,0,0,0,0])

            elif(CompareWithTolerance(10, row["VA"], 1) and
                CompareWithTolerance(10, row["VB"], 1) and
                CompareWithTolerance(10, row["VC"], 1) and
                CompareWithTolerance(6, row["IA"], 1) and
                CompareWithTolerance(6, row["IB"], 1) and
                CompareWithTolerance(6, row["IC"], 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDA"]), 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDB"]), 1) and
                CompareWithTolerance(expectedVal, toPhase(row["PFDC"]), 1)
                ):
                if(i1to3pop):
                    tests.append(["P6A4to6",expectedVal, 0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,toPhase(row["PFDA"]), toPhase(row["PFDB"]),toPhase(row["PFDC"]),0,0,0,0,0,0,0,0,0,0,0,0])
                else:
                    tests.append(["P6A1to3",expectedVal,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,toPhase(row["PFDA"]), toPhase(row["PFDB"]),toPhase(row["PFDC"]),0,0,0,0,0,0,0,0,0,0,0,0,0,0,0])
                if(CompareWithTolerance(180, toPhase(row["PFDA"]), 1)):
                    seti1to3True()

    if(CompareWithTolerance(15, row["VA"], 1) and
        CompareWithTolerance(10, row["VB"], 1) and
        CompareWithTolerance(15, row["VC"], 1) and
        CompareWithTolerance(1.5, row["IA"], 1) and
        CompareWithTolerance(1, row["IB"], 1) and
        CompareWithTolerance(1.5, row["IC"], 1) and
        CompareWithTolerance(1, row["PFDA"], .1) and
        CompareWithTolerance(1, row["PFDB"], .1) and
        CompareWithTolerance(1, row["PFDC"], .1)
        ):
        setvlowcompletetrue()
                    
    if(vlowcomplete):
        for expectedVal in expectedVals["VH"]:
            if(CompareWithTolerance(expectedVal, row["VA"], 1) and
                CompareWithTolerance(expectedVal, row["VB"], 1) and
                CompareWithTolerance(expectedVal, row["VC"], 1) and
                CompareWithTolerance(0, row["IA"], 1) and
                CompareWithTolerance(0, row["IB"], 1) and
                CompareWithTolerance(0, row["IC"], 1) and
                CompareWithTolerance(1, row["PFDA"], .1) and
                CompareWithTolerance(1, row["PFDB"], .1) and
                CompareWithTolerance(1, row["PFDC"], .1)
                ):
                
                    tests.append(["VH1to3",expectedVal,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,row["VA"], row["VB"],row["VC"],0,0,0,0,0,0,0,0,0])

        for expectedPhaseVal in expectedVals["PFDA"]:
            for expectedVal in phaseTests["HVPT"]:
                if( CompareWithTolerance(expectedVal[1], row["VA"], 1) and
                    CompareWithTolerance(expectedVal[1], row["VB"], 1) and
                    CompareWithTolerance(expectedVal[1], row["VC"], 1) and
                    CompareWithTolerance(1, row["IA"], 1) and
                    CompareWithTolerance(1, row["IB"], 1) and
                    CompareWithTolerance(1, row["IC"], 1) and
                    CompareWithTolerance(expectedPhaseVal, toPhase(row["PFDA"]), 1) and
                    CompareWithTolerance(expectedPhaseVal, toPhase(row["PFDB"]), 1) and
                    CompareWithTolerance(expectedPhaseVal, toPhase(row["PFDC"]), 1)
                    ):
                    tests.append([expectedVal[0],expectedPhaseVal,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,toPhase(row["PFDA"]), toPhase(row["PFDB"]),toPhase(row["PFDC"]),0,0,0,0,0,0])
        
        for expectedVal in expectedVals["IH"]:
            if( CompareWithTolerance(0, row["VA"], 1) and
                CompareWithTolerance(0, row["VB"], 1) and
                CompareWithTolerance(0, row["VC"], 1) and
                CompareWithTolerance(expectedVal, row["IA"], 1) and
                CompareWithTolerance(expectedVal, row["IB"], 1) and
                CompareWithTolerance(expectedVal, row["IC"], 1) and
                CompareWithTolerance(1, row["PFDA"], .1) and
                CompareWithTolerance(1, row["PFDB"], .1) and
                CompareWithTolerance(1, row["PFDC"], .1)
                ):
                
                    tests.append(["IH1to3",expectedVal,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,row["IA"], row["IB"],row["IC"],0,0,0])

        for expectedPhaseVal in expectedVals["PFDA"]:
            for expectedVal in phaseTests["HIPT"]:
                if( CompareWithTolerance(10, row["VA"], 1) and
                    CompareWithTolerance(10, row["VB"], 1) and
                    CompareWithTolerance(10, row["VC"], 1) and
                    CompareWithTolerance(expectedVal[1], row["IA"], 1) and
                    CompareWithTolerance(expectedVal[1], row["IB"], 1) and
                    CompareWithTolerance(expectedVal[1], row["IC"], 1) and
                    CompareWithTolerance(expectedPhaseVal, toPhase(row["PFDA"]), 1) and
                    CompareWithTolerance(expectedPhaseVal, toPhase(row["PFDB"]), 1) and
                    CompareWithTolerance(expectedPhaseVal, toPhase(row["PFDC"]), 1)
                    ):
                    tests.append([expectedVal[0],expectedPhaseVal,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,toPhase(row["PFDA"]), toPhase(row["PFDB"]),toPhase(row["PFDC"])])

    return row


df['Time'] = df['Time'].map(lambda x: x.replace(" ", ""))
pd.to_datetime(df['Time'], format='%H:%M:%S')
df.set_index('Time', inplace=True)

df.apply(lambda x: FilterAndIdentify(x), axis = 1)
testdf = pd.DataFrame(tests, columns = testcols)
testdf.to_excel(os.path.dirname(os.path.realpath(__file__)) + "\\data\\.FilteredData.xlsx")

templatedf = pd.read_excel(os.path.dirname(os.path.realpath(__file__)) + "\\data\\DobleF6150.xlsx", )
currentItem = None

wb2 = load_workbook(os.path.dirname(os.path.realpath(__file__)) + "\\data\\DobleF6150.xlsx")

templatews1 = wb2.worksheets[0]

def SetCurrentItem(val):
    global currentItem
    currentItem = val

def GetTestResults(row):
    if( isinstance(row["Item"], str)):
        SetCurrentItem(row["Item"])
    else:
        print(type(row["Item"]))
    tempdf = testdf[(testdf == float(row["Level"]) ).any(axis = 1)]
    tempdf = tempdf.filter(items = [currentItem])
    tempdf = tempdf[(tempdf != 0).any(axis = 1)]
    tempdf["Dif"] = np.absolute(tempdf[currentItem] - row["Level"])
    maxdev = tempdf.iloc[tempdf['Dif'].argmax(), 0]
    results.append(maxdev)
    print(tempdf.head(5))

results = []

templatedf.apply(lambda x: GetTestResults(x), axis = 1)
i = 2
for item in results:
    templatews1.cell(i, 5).value = item
    i = i+1

wb2.save(os.path.dirname(os.path.realpath(__file__)) + "\\Results.xlsx")


